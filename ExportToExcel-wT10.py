import openpyxl

# Read the text file
with open('result.txt', 'r') as f:
    data = f.readlines()

# Create a new workbook
workbook = openpyxl.Workbook()
sheet = workbook.active

# Write the data to the worksheet
for i, line in enumerate(data):
    line = eval(line.strip())
    sheet.cell(row=i+1, column=1, value=line['first_name'])
    sheet.cell(row=i+1, column=2, value=line['last_name'])
    sheet.cell(row=i+1, column=3, value=line['email'])

# Save the workbook
workbook.save('ExportedForm.xlsx')
